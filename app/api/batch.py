from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any
from uuid import uuid4
import asyncio
import json
from app.core.config import settings

router = APIRouter()

import redis
try:
    redis_client = redis.Redis.from_url(settings.redis_url, decode_responses=True)
    # Check if redis connects
    redis_client.ping()
except Exception:
    redis_client = None

batch_results: dict = {}

async def _process_batch(batch_id: str, payloads: List[Dict]):
    from app.validation.validator import InvoiceValidator
    from app.adapters.generic_erp import GenericJSONAdapter
    
    validator = InvoiceValidator()
    adapter = GenericJSONAdapter()
    
    results = []
    for i, payload in enumerate(payloads):
        try:
            invoice = adapter.transform(payload)
            report = validator.validate(invoice)
            results.append({"invoice_number": payload.get("invoice_number"), "is_valid": report.is_valid, "errors": [e.__dict__ for e in report.errors]})
        except Exception as e:
            results.append({"invoice_number": payload.get("invoice_number"), "is_valid": False, "error": str(e)})
            
        data = {"status": "PROCESSING", "total": len(payloads), "done": i + 1, "results": results if (i + 1) == len(payloads) else []}
        if (i + 1) == len(payloads):
             data["status"] = "COMPLETE"
             
        if redis_client:
            redis_client.setex(f"batch:{batch_id}", 3600, json.dumps(data))
        else:
            batch_results[batch_id] = data

@router.post("/batch-validate")
async def batch_validate(payloads: List[Dict[str, Any]], background_tasks: BackgroundTasks):
    if len(payloads) > 500:
        raise HTTPException(status_code=400, detail="Batch size limit is 500 invoices")
    batch_id = f"BATCH-{uuid4().hex[:8].upper()}"
    
    initial_data = {"status": "PROCESSING", "total": len(payloads), "done": 0}
    if redis_client:
        redis_client.setex(f"batch:{batch_id}", 3600, json.dumps(initial_data))
    else:
        batch_results[batch_id] = initial_data
        
    background_tasks.add_task(_process_batch, batch_id, payloads)
    return {"batch_id": batch_id, "status": "ACCEPTED", "total": len(payloads), "poll_url": f"/api/v1/batch-status/{batch_id}"}

def map_excel_to_pint_ae(df) -> List[Dict]:
    """
    Basic mapping, filling NaN with empty strings and converting to dict.
    The actual deep transformation to InvoicePayload is handled by the Adapter layer.
    """
    df = df.fillna("")
    return df.to_dict("records")

@router.post("/upload-excel")
async def upload_excel(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(400, "Only Excel/CSV files accepted")
    
    contents = await file.read()
    batch_id = f"BATCH-{str(uuid4().hex[:8]).upper()}"
    
    import pandas as pd
    import io
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(contents), dtype=str)
    except Exception as e:
        raise HTTPException(400, f"Error parsing file: {str(e)}")
        
    payloads = map_excel_to_pint_ae(df)
    
    initial_data = {"status": "PROCESSING", "total": len(payloads), "done": 0}
    if redis_client:
        redis_client.setex(f"batch:{batch_id}", 3600, json.dumps(initial_data))
    else:
        batch_results[batch_id] = initial_data
        
    background_tasks.add_task(_process_batch, batch_id, payloads)
    return {"batch_id": batch_id, "total_rows": len(payloads), "status": "PROCESSING"}

@router.get("/download-template")
async def download_template():
    """
    Returns a professional Excel template for bulk uploads with sample PINT AE data.
    Pre-formats specific columns as Text to prevent integer-casting issues.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    TEXT_COLUMNS = {
        "invoice_type_code", "payment_means_type_code", "currency_code",
        "tax_category_code", "transaction_type", "seller_country_code",
        "buyer_country_code", "unit_of_measure", "tax_category",
        "seller_trn", "buyer_trn", "line_id",
    }
    
    TEMPLATE_COLUMNS = [
        "invoice_number", "invoice_date", "invoice_type_code",
        "payment_means_type_code", "transaction_type", "currency_code",
        "tax_category_code",
        "seller_name", "seller_trn", "seller_address", "seller_country_code",
        "buyer_name", "buyer_trn", "buyer_address", "buyer_country_code",
        "line_id", "item_name", "unit_of_measure",
        "quantity", "unit_price", "line_net_amount",
        "tax_category", "tax_rate", "tax_amount",
        "total_without_tax", "tax_amount", "total_with_tax", "amount_due",
    ]
    
    SAMPLE_ROWS = [
        {
            "invoice_number": "INV-2026-001",
            "invoice_date": "2026-04-01",
            "invoice_type_code": "380",        
            "payment_means_type_code": "10",   
            "transaction_type": "B2B",
            "currency_code": "AED",
            "tax_category_code": "S",
            "seller_name": "Adamas Tech Corp",
            "seller_trn": "100200300400500",
            "seller_address": "Dubai, UAE",
            "seller_country_code": "AE",
            "buyer_name": "Client Group FZE",
            "buyer_trn": "100999888777666",
            "buyer_address": "Abu Dhabi, UAE",
            "buyer_country_code": "AE",
            "line_id": "1",
            "item_name": "Consulting Services",
            "unit_of_measure": "EA",
            "quantity": 10,
            "unit_price": 500.00,
            "line_net_amount": 5000.00,
            "tax_category": "S",
            "tax_rate": 0.05,
            "tax_amount": 250.00,
            "total_without_tax": 5000.00,
            "total_with_tax": 5250.00,
            "amount_due": 5250.00,
        },
        {
            "invoice_number": "INV-2026-002",
            "invoice_date": "2026-04-01",
            "invoice_type_code": "380",
            "payment_means_type_code": "30",   
            "transaction_type": "B2C",
            "currency_code": "AED",
            "tax_category_code": "S",
            "seller_name": "Adamas Tech Corp",
            "seller_trn": "100200300400500",
            "seller_address": "Dubai, UAE",
            "seller_country_code": "AE",
            "buyer_name": "Individual Customer",
            "buyer_trn": "",                   
            "buyer_address": "Sharjah, UAE",
            "buyer_country_code": "AE",
            "line_id": "1",
            "item_name": "Software License",
            "unit_of_measure": "EA",
            "quantity": 2,
            "unit_price": 100.00,
            "line_net_amount": 200.00,
            "tax_category": "S",
            "tax_rate": 0.05,
            "tax_amount": 10.00,
            "total_without_tax": 200.00,
            "total_with_tax": 210.00,
            "amount_due": 210.00,
        },
    ]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "PINT_AE_Invoices"
    
    header_fill = PatternFill("solid", fgColor="1a6fcf")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 16)
        
        if col_name in TEXT_COLUMNS:
            for row_idx in range(2, 502):   
                ws.cell(row=row_idx, column=col_idx).number_format = "@"
                
    for row_idx, row_data in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_name in TEXT_COLUMNS:
                cell.number_format = "@"
                cell.value = str(val) if val != "" else ""
                
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=PINT_AE_Bulk_Template.xlsx"}
    )

@router.get("/batch-status/{batch_id}")
async def batch_status(batch_id: str):
    if redis_client:
        data = redis_client.get(f"batch:{batch_id}")
        if data:
            return json.loads(data)
    elif batch_id in batch_results:
        return batch_results[batch_id]
        
    raise HTTPException(status_code=404, detail="Batch not found")
